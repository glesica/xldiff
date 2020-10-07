import sys

from openpyxl import load_workbook

from xldiff.diff import first_difference, print_row


def diff_files(left_filename, right_filename):
    left = load_workbook(left_filename, data_only=True)
    right = load_workbook(right_filename, data_only=True)
    diff, index = first_difference(left.active, right.active)

    if diff is None:
        sys.exit(0)

    print(f"Difference at row {index}")

    print("⬅ ", end="")
    print_row(diff[0])

    print("➡ ", end="")
    print_row(diff[1])

    sys.exit(1)
